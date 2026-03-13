import os
import warnings
import numpy as np
import pandas as pd
from itertools import product
from multiprocessing.pool import ThreadPool
from multiprocessing import cpu_count
from sklearn.model_selection import KFold
from sklearn.preprocessing import StandardScaler
from sksurv.linear_model import CoxnetSurvivalAnalysis
from sksurv.util import Surv

from machine_learning.utils.data.clean_survival_inputs import clean_survival_inputs


# ── module-level worker ─────────────────────────────────────────────────────
# Must be module-level so multiprocessing.Pool can pickle it.

def _evaluate_params(args):
    """
    Evaluate one (alpha, l1_ratio) combination across all CV folds.

    Returns
    -------
    combo : (alpha, l1_ratio)
        Echoed back so imap_unordered stays self-contained.
    result : (mean_c_index | None, params_dict | None, diag_dict)
    """
    df_fit, alpha, l1_ratio, kf_splits = args

    feature_cols = [c for c in df_fit.columns if c not in ("T", "E")]
    X_all = df_fit[feature_cols].values.astype(float)
    y_all = Surv.from_arrays(
        event=df_fit["E"].astype(bool).values,
        time=df_fit["T"].astype(float).values,
    )

    c_indices         = []
    n_folds_failed    = 0
    fit_exception_msg = ""

    for train_idx, val_idx in kf_splits:
        X_train_raw, X_val_raw = X_all[train_idx], X_all[val_idx]
        y_train,     y_val     = y_all[train_idx], y_all[val_idx]

        # Scale on train, apply to val — mirrors _safe_scale in generate_survival_features
        scaler  = StandardScaler()
        X_train = np.clip(scaler.fit_transform(X_train_raw), -10, 10)
        X_val   = np.clip(scaler.transform(X_val_raw),       -10, 10)

        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                cox = CoxnetSurvivalAnalysis(
                    l1_ratio=l1_ratio,
                    alphas=[alpha],
                    fit_baseline_model=False,  # C-index only — baseline not needed here
                    max_iter=100_000,
                    tol=1e-7,
                )
                cox.fit(X_train, y_train)

            c_index = cox.score(X_val, y_val)
            c_indices.append(float(c_index))

        except Exception as exc:
            n_folds_failed += 1
            fit_exception_msg = str(exc)[:120]

    diag = {
        "n_folds_ok":     len(c_indices),
        "n_folds_failed": n_folds_failed,
        "c_index_std":    float(np.std(c_indices)) if len(c_indices) > 1 else np.nan,
        "fit_exception":  fit_exception_msg,
    }

    if not c_indices:
        return (alpha, l1_ratio), (None, None, diag)

    return (alpha, l1_ratio), (
        float(np.mean(c_indices)),
        {"alpha": alpha, "l1_ratio": l1_ratio},
        diag,
    )


# ── main class ──────────────────────────────────────────────────────────────

class CoxHyperparameterTuner:
    """
    Tunes CoxnetSurvivalAnalysis hyperparameters (alpha x l1_ratio) via
    parallelised cross-validation and produces an optional Excel report.

    Usage
    -----
    tuner = CoxHyperparameterTuner(save_report_path="results/")
    tuner.fit(X_surv, T, E)

    print(tuner.best_params_)     # {"alpha": float, "l1_ratio": float}
    print(tuner.best_c_index_)
    print(tuner.results_df_)

    Parameters
    ----------
    alpha_grid : list of float
        Regularisation strengths. Smaller = less regularisation.
        Maps to the old "penalizer" parameter.
    l1_ratios : list of float
        Elastic-net mixing. 0 = pure Ridge, 1 = pure Lasso.
    n_splits : int
        K-Fold CV splits.
    random_state : int
    n_jobs : int
        Worker processes. -1 = all cores.
    save_report_path : str or None
        Path for the Excel diagnostics report.
    """

    def __init__(
        self,
        alpha_grid       = [0.001, 0.01, 0.1, 1, 10, 100],
        l1_ratios        = [0, 0.25, 0.5, 0.75, 1],
        n_splits         = 5,
        random_state     = 42,
        n_jobs           = -1,
        save_report_path = None,
    ):
        self.alpha_grid       = alpha_grid
        self.l1_ratios        = l1_ratios
        self.n_splits         = n_splits
        self.random_state     = random_state
        self.n_jobs           = n_jobs
        self.save_report_path = save_report_path

        self.best_params_  = None
        self.best_c_index_ = None
        self.results_df_   = None
        self._df_fit       = None
        self._kf_splits    = None

    # ── public API ───────────────────────────────────────────────────────────

    def fit(self, X_surv, T, E):
        """
        Run the full tuning pipeline.

        Returns self (allows chaining: tuner.fit(X, T, E).best_params_)
        """
        self._df_fit    = self._clean_data(X_surv, T, E)
        self._kf_splits = self._build_folds()
        combos          = self._build_grid()
        self._run_search(combos)

        if self.save_report_path:
            self._save_report()

        return self

    # ── pipeline steps ───────────────────────────────────────────────────────

    def _clean_data(self, X_surv, T, E) -> pd.DataFrame:
        _, _, _, df_fit = clean_survival_inputs(X_surv, T, E)
        return df_fit

    def _build_folds(self) -> list:
        kf = KFold(n_splits=self.n_splits, shuffle=True, random_state=self.random_state)
        return list(kf.split(self._df_fit))

    def _build_grid(self) -> list:
        combos = list(product(self.alpha_grid, self.l1_ratios))
        print(
            f"[CoxHyperparameterTuner] Grid: {len(combos)} combinations "
            f"x {self.n_splits} folds | workers={self.n_jobs}",
            flush=True,
        )
        return combos

    def _run_search(self, combos: list) -> None:
        n_workers = cpu_count() if self.n_jobs == -1 else self.n_jobs
        total     = len(combos)

        task_args = [
            (self._df_fit, alpha, l1_ratio, self._kf_splits)
            for alpha, l1_ratio in combos
        ]

        raw_results = []
        with ThreadPool(processes=cpu_count()) as pool:
            for i, result in enumerate(
                pool.imap_unordered(_evaluate_params, task_args), start=1
            ):
                raw_results.append(result)
                if i % 10 == 0 or i == total:
                    print(
                        f"[CoxHyperparameterTuner] {i}/{total} combos evaluated...",
                        flush=True,
                    )

        rows         = []
        best_c_index = -np.inf
        best_params  = None

        for (alpha, l1_ratio), (mean_c, params, diag) in raw_results:
            rows.append({
                "alpha":        alpha,
                "l1_ratio":     l1_ratio,
                "c_index_mean": mean_c if mean_c is not None else np.nan,
                **diag,
            })
            if mean_c is not None and mean_c > best_c_index:
                best_c_index = mean_c
                best_params  = params

        if best_params is None:
            print(
                "[CoxHyperparameterTuner] No valid params — falling back to defaults.",
                flush=True,
            )
            best_params  = {"alpha": 0.1, "l1_ratio": 1.0}
            best_c_index = np.nan

        self.results_df_   = (
            pd.DataFrame(rows)
            .sort_values("c_index_mean", ascending=False)
            .reset_index(drop=True)
        )
        self.best_params_  = best_params
        self.best_c_index_ = best_c_index

        print(
            f"[CoxHyperparameterTuner] Best -> {self.best_params_} | "
            f"C-index: {self.best_c_index_:.4f}",
            flush=True,
        )

    def _save_report(self) -> None:
        path = self.save_report_path
        if os.path.isdir(path) or path.endswith(("/", "\\")):
            path = os.path.join(path, "cox_tuning_report.xlsx")
        elif not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
        _build_excel_report(self.results_df_, path)

    # ── convenience ──────────────────────────────────────────────────────────

    def _check_is_fitted(self):
        if self.best_params_ is None:
            raise RuntimeError("Call fit() before accessing results.")

    @property
    def summary(self) -> pd.DataFrame:
        """Top-10 results sorted by C-index."""
        self._check_is_fitted()
        return self.results_df_.head(10)


# ── Excel report ─────────────────────────────────────────────────────────────

def _health_flag(row) -> str:
    flags = []
    if row["n_folds_failed"] > 0:
        flags.append(f"fold_fail x{row['n_folds_failed']}")
    if row.get("fit_exception", ""):
        flags.append("exception")
    return "; ".join(flags) if flags else "OK"


def _build_excel_report(df: pd.DataFrame, save_path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule

    df = df.copy()
    df["health_flag"] = df.apply(_health_flag, axis=1)

    wb = Workbook()

    thin_side   = Side(style="thin", color="CCCCCC")
    thin_border = Border(left=thin_side, right=thin_side,
                         top=thin_side,  bottom=thin_side)
    center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    best_fill = PatternFill("solid", fgColor="BDD7EE")
    alt_fill  = PatternFill("solid", fgColor="F5F5F5")
    ok_fill   = PatternFill("solid", fgColor="E2EFDA")
    warn_fill = PatternFill("solid", fgColor="FFEB9C")
    bad_fill  = PatternFill("solid", fgColor="FFC7CE")
    sec_fill  = PatternFill("solid", fgColor="D6E4F0")

    # ── Sheet 1: All Results ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "All Results"

    HEADERS    = ["Rank", "alpha", "l1_ratio", "c_index_mean", "c_index_std",
                  "folds_ok", "folds_failed", "fit_exception", "health_flag"]
    COL_WIDTHS = [6, 12, 10, 14, 12, 10, 12, 42, 24]

    for ci, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = center; cell.border = thin_border
    ws.row_dimensions[1].height = 30

    for ri, row in df.iterrows():
        er   = ri + 2
        rank = ri + 1
        flag = row["health_flag"]

        vals = [
            rank,
            row["alpha"], row["l1_ratio"],
            round(row["c_index_mean"], 6) if pd.notna(row["c_index_mean"]) else "FAILED",
            round(row["c_index_std"],  6) if pd.notna(row["c_index_std"])  else "",
            row["n_folds_ok"], row["n_folds_failed"],
            row["fit_exception"] or "", flag,
        ]

        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=er, column=ci, value=val)
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin_border
            cell.fill      = best_fill if rank == 1 else (
                alt_fill if ri % 2 == 0 else PatternFill()
            )

        fc = ws.cell(row=er, column=len(HEADERS))
        if flag == "OK":
            fc.fill = ok_fill;   fc.font = Font(name="Arial", size=9, color="375623")
        elif "exception" in flag or "fold_fail" in flag:
            fc.fill = bad_fill;  fc.font = Font(name="Arial", size=9, color="9C0006")
        else:
            fc.fill = warn_fill; fc.font = Font(name="Arial", size=9, color="7D4900")

    n_data = len(df) + 1
    cc = get_column_letter(4)   # c_index_mean column
    ws.conditional_formatting.add(
        f"{cc}2:{cc}{n_data}",
        ColorScaleRule(
            start_type="min",  start_color="FFC7CE",
            mid_type="num",    mid_value=0.6, mid_color="FFEB9C",
            end_type="max",    end_color="C6EFCE",
        ),
    )
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # ── Sheet 2: Best Result ─────────────────────────────────────────────────
    ws2  = wb.create_sheet("Best Result")
    best = df.iloc[0]
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 28

    ws2.merge_cells("A1:B1")
    ws2["A1"] = "CoxNet Tuning — Best Result"
    ws2["A1"].font      = Font(name="Arial", bold=True, size=13, color="1F4E79")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    summary = [
        ("Parameter",         "Value"),
        ("alpha",             best["alpha"]),
        ("l1_ratio",          best["l1_ratio"]),
        ("", ""),
        ("Metric",            "Value"),
        ("C-index (mean CV)", round(best["c_index_mean"], 6)),
        ("C-index (std  CV)", round(best["c_index_std"],  6) if pd.notna(best["c_index_std"]) else "N/A"),
        ("Folds OK",          best["n_folds_ok"]),
        ("Folds Failed",      best["n_folds_failed"]),
        ("", ""),
        ("Diagnostics",       "Value"),
        ("Health Flag",       best["health_flag"]),
        ("Fit Exception",     best["fit_exception"] or "None"),
    ]
    section_rows = {2, 6, 11}
    for ri, (lbl, val) in enumerate(summary, 2):
        a = ws2.cell(row=ri, column=1, value=lbl)
        b = ws2.cell(row=ri, column=2, value=val)
        if ri in section_rows:
            for c in (a, b):
                c.fill = sec_fill
                c.font = Font(name="Arial", bold=True, size=10, color="1F4E79")
        else:
            a.font = Font(name="Arial", bold=True, size=10)
            b.font = Font(name="Arial", size=10)
        for c in (a, b):
            c.alignment = Alignment(vertical="center")
            c.border    = thin_border

    # ── Sheet 3: Diagnostics Overview ────────────────────────────────────────
    ws3 = wb.create_sheet("Diagnostics Overview")
    ws3.column_dimensions["A"].width = 36
    ws3.column_dimensions["B"].width = 20

    ws3.merge_cells("A1:B1")
    ws3["A1"] = "Diagnostics Overview"
    ws3["A1"].font      = Font(name="Arial", bold=True, size=13, color="1F4E79")
    ws3["A1"].alignment = Alignment(horizontal="center")
    ws3.row_dimensions[1].height = 28

    total = len(df)
    diag_summary = [
        ("Metric",                           "Count"),
        ("Total combinations evaluated",     total),
        ("Combinations — no issues (OK)",    (df["health_flag"] == "OK").sum()),
        ("Combinations — failed folds",      (df["n_folds_failed"] > 0).sum()),
        ("Combinations — exceptions",        (df["fit_exception"].astype(str).str.len() > 0).sum()),
        ("", ""),
        ("Best  C-index (mean CV)",          round(df["c_index_mean"].max(),    6)),
        ("Worst C-index (mean CV)",          round(df["c_index_mean"].min(),    6)),
        ("Median C-index (mean CV)",         round(df["c_index_mean"].median(), 6)),
        ("Mean  C-index std across combos",  round(df["c_index_std"].mean(),    6)),
    ]
    for ri, (lbl, val) in enumerate(diag_summary, 2):
        a = ws3.cell(row=ri, column=1, value=lbl)
        b = ws3.cell(row=ri, column=2, value=val)
        if ri == 2:
            for c in (a, b):
                c.fill = sec_fill
                c.font = Font(name="Arial", bold=True, size=10, color="1F4E79")
        else:
            a.font = Font(name="Arial", bold=True, size=10)
            b.font = Font(name="Arial", size=10)
        for c in (a, b):
            c.alignment = Alignment(vertical="center")
            c.border    = thin_border

    wb.save(save_path)
    print(f"[CoxHyperparameterTuner] Report saved -> {save_path}", flush=True)